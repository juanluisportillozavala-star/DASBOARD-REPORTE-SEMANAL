"""
=========================================================
PROCESAMIENTO DEL MÓDULO INVENTARIO  (v2: cruce de 2 BD)
=========================================================
Ahora el inventario se alimenta de DOS archivos que se cruzan
por el CÓDIGO de producto (ej. "LD008-2"):

  1) stock_valuation_layer (tabla dinámica de Odoo):
     aporta la FECHA de entrada real de cada producto.
     - Un producto puede tener varias entradas (varias fechas);
       se toma la MÁS ANTIGUA para medir lento movimiento.
     - Cantidad y Valor se SUMAN de todas sus entradas.

  2) Quants (stock_quant): aporta la UBICACIÓN de cada producto.

Regla: los productos que NO tengan ubicación en Quants se
DESCARTAN (no manejamos esas ubicaciones).

La fecha para "DIAS EN ALMACEN" es una fecha de corte MANUAL
que se captura al cargar (como en Cartera).

Salida: mismas columnas que la versión anterior, para que la
tabla, KPIs y gráficos no cambien:
  Ubicación, Producto, Unidad de medida, Cantidad en inventario,
  Fecha de entrada, Valor, DIAS EN ALMACEN, CATEGORIA
"""

import base64
import io
import re
import pandas as pd
import numpy as np

# columnas de salida (compatibles con tabla_inventario.py)
COL_PRODUCTO = "Producto"
COL_UBICACION = "Ubicación"
COL_UNIDAD = "Unidad de medida"
COL_CANTIDAD = "Cantidad en inventario"
COL_FECHA_ENTRADA = "Fecha de entrada"
COL_VALOR = "Valor"

CAT_1 = "1-30 días"
CAT_2 = "31-60 días"
CAT_3 = "61+ días"

_MESES_ES = {"ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
             "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12}


def _codigo(txt):
    """Extrae el código entre corchetes: '[LD008-2] ...' -> 'LD008-2'."""
    if txt is None or (isinstance(txt, float) and pd.isna(txt)):
        return None
    m = re.match(r"\s*\[([^\]]+)\]", str(txt))
    return m.group(1) if m else None


def _parse_fecha_es(s):
    """Convierte '21 feb. 2024' a Timestamp."""
    m = re.match(r"(\d+)\s+(\w+)\.?\s+(\d+)", str(s))
    if not m:
        return None
    dia, mes, anio = int(m.group(1)), m.group(2)[:3].lower(), int(m.group(3))
    mnum = _MESES_ES.get(mes)
    if not mnum:
        return None
    return pd.Timestamp(year=anio, month=mnum, day=dia)


def _b64_a_bytes(contents):
    if contents is None:
        return None
    return io.BytesIO(base64.b64decode(contents.split(",")[1]))


def _leer_valuation(contents):
    """Lee la tabla dinámica stock_valuation_layer y devuelve, por
    producto: código, producto, fecha_mas_antigua, cantidad, valor."""
    import openpyxl
    wb = openpyxl.load_workbook(_b64_a_bytes(contents), data_only=True)
    ws = wb[wb.sheetnames[0]]

    # mapear columnas de "Cant. restante" a su fecha (fila 2 = fecha,
    # fila 3 = tipo). Cada fecha ocupa un par (cant, valor).
    col_fecha = {}
    for c in range(2, ws.max_column + 1):
        f = ws.cell(row=2, column=c).value
        tipo = ws.cell(row=3, column=c).value
        if f and f != "Total" and tipo == "Cant. restante":
            fecha = _parse_fecha_es(f)
            if fecha is not None:
                col_fecha[c] = fecha

    registros = []
    for r in range(5, ws.max_row + 1):
        prod = ws.cell(row=r, column=1).value
        cod = _codigo(prod)
        if not cod:
            continue
        fechas, cant_tot, val_tot = [], 0.0, 0.0
        for c, fecha in col_fecha.items():
            cant = ws.cell(row=r, column=c).value
            valor = ws.cell(row=r, column=c + 1).value
            if cant not in (None, 0):
                fechas.append(fecha)
                cant_tot += float(cant or 0)
                val_tot += float(valor or 0)
        if fechas:
            registros.append({
                "codigo": cod,
                "producto": str(prod).strip(),
                "fecha_entrada": min(fechas),   # la MÁS ANTIGUA
                "cantidad": cant_tot,
                "valor": val_tot,
            })
    wb.close()
    return pd.DataFrame(registros)


def _leer_quants(contents):
    """Lee Quants y devuelve código -> (ubicación, unidad)."""
    df = pd.read_excel(_b64_a_bytes(contents))
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(subset=[COL_PRODUCTO])
    df = df[~df[COL_PRODUCTO].astype(str).str.contains("Existencias", case=False, na=False)]
    df["codigo"] = df[COL_PRODUCTO].apply(_codigo)
    df["ubicacion"] = df[COL_UBICACION].astype(str).str.split("/").str[0]
    unidad = df.groupby("codigo")[COL_UNIDAD].first().to_dict() if COL_UNIDAD in df.columns else {}
    ubic = df.dropna(subset=["codigo"]).groupby("codigo")["ubicacion"].first().to_dict()
    return ubic, unidad


def procesar_bd_inventario(contents_valuation, contents_quants, fecha_corte=None):
    """Cruza las dos BD y arma el DataFrame final del módulo.
    fecha_corte: fecha manual para calcular DIAS EN ALMACEN."""
    if fecha_corte is None:
        fecha_corte = pd.Timestamp.today().normalize()
    else:
        fecha_corte = pd.Timestamp(fecha_corte).normalize()

    val = _leer_valuation(contents_valuation)
    if val is None or len(val) == 0:
        raise Exception("No se pudieron leer productos del archivo de valuación.")

    mapa_ubic, mapa_unidad = _leer_quants(contents_quants)

    # cruzar: pegar ubicación; descartar los que no tengan
    val["Ubicación"] = val["codigo"].map(mapa_ubic)
    val = val[val["Ubicación"].notna()].copy()
    val["Unidad de medida"] = val["codigo"].map(mapa_unidad)

    # días en almacén (fecha de entrada normalizada, como en v1)
    entrada = pd.to_datetime(val["fecha_entrada"], errors="coerce").dt.normalize()
    val["DIAS EN ALMACEN"] = (fecha_corte - entrada).dt.days

    val["CATEGORIA"] = pd.cut(
        val["DIAS EN ALMACEN"], bins=[0, 30, 60, 999999],
        labels=[CAT_1, CAT_2, CAT_3],
    ).astype(str)

    # armar salida con los nombres de columna que espera la tabla
    out = pd.DataFrame({
        COL_UBICACION: val["Ubicación"],
        COL_PRODUCTO: val["producto"],
        COL_UNIDAD: val["Unidad de medida"],
        COL_CANTIDAD: val["cantidad"],
        COL_FECHA_ENTRADA: entrada,
        COL_VALOR: val["valor"],
        "DIAS EN ALMACEN": val["DIAS EN ALMACEN"],
        "CATEGORIA": val["CATEGORIA"],
    }).reset_index(drop=True)

    return out


def leer_archivo(contents_valuation, contents_quants, fecha_corte=None):
    return procesar_bd_inventario(contents_valuation, contents_quants, fecha_corte)