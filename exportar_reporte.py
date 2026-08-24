"""
=========================================================
exportar_reporte.py  —  Descargar el REPORTE en Excel con las
tablas dinámicas VIVAS
=========================================================
Toma la plantilla (plantillas/REPORTE_SEMANAL_FINAL.xlsx), que
ya trae las 7 tablas dinámicas y los segmentadores, y reemplaza
SOLO los datos de las hojas BD (BD Ventas, BD Cobranza, BD
Cartera, BD CxP) con la BD CRUDA guardada en Supabase.

Técnica: cirugía de ZIP/XML (openpyxl NO soporta pivots/slicers
y los rompería). Se reescribe únicamente el sheetN.xml de cada
hoja BD; todo lo demás (pivots, caches, slicers) se copia intacto.

Además:
  • Se preserva el estilo (formato) de cada columna leyéndolo de
    la fila 2 original -> las fechas siguen viéndose como fechas.
  • Se ajusta el rango de origen de cada pivotCache al nuevo
    número de filas.
  • Se activa refreshOnLoad=1 -> las dinámicas se refrescan solas
    al abrir el archivo.

Requiere que la BD cruda esté guardada (db.leer_crudo). Si un
módulo no tiene BD cruda aún, esa hoja se deja como está en la
plantilla.
"""

import io
import re
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime, date
import pandas as pd
from openpyxl.utils import get_column_letter

import db

# módulo interno -> nombre de la hoja BD en la plantilla
MODULO_HOJA = {
    "ventas": "BD Ventas",
    "ingresos": "BD Cobranza",
    "cartera": "BD Cartera",
    "saldo_proveedor": "BD CxP",
}

# módulo interno -> nombre de la hoja de la TABLA DINÁMICA (la que
# se deja visible cuando se descarga "solo ese módulo").
MODULO_PIVOTE = {
    "ventas": "Ventas",
    "ingresos": "Ingreso",
    "cartera": "Cartera",
    "saldo_proveedor": "Saldo Prov",
}

_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


# ---------------------------------------------------------
# helpers de celda
# ---------------------------------------------------------

def _es_fecha(v):
    return isinstance(v, (datetime, date, pd.Timestamp))


def _excel_serial(v):
    if isinstance(v, pd.Timestamp):
        v = v.to_pydatetime()
    if isinstance(v, date) and not isinstance(v, datetime):
        v = datetime(v.year, v.month, v.day)
    base = datetime(1899, 12, 30)
    delta = v - base
    return delta.days + delta.seconds / 86400.0


def _celda_xml(ref, valor, estilo):
    s_attr = f' s="{estilo}"' if estilo else ""
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return f'<c r="{ref}"{s_attr}/>'
    if _es_fecha(valor):
        return f'<c r="{ref}"{s_attr}><v>{_excel_serial(valor)}</v></c>'
    if isinstance(valor, bool):
        return f'<c r="{ref}" t="b"{s_attr}><v>{1 if valor else 0}</v></c>'
    if isinstance(valor, (int, float)):
        return f'<c r="{ref}"{s_attr}><v>{valor}</v></c>'
    txt = (str(valor).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))
    return (f'<c r="{ref}" t="inlineStr"{s_attr}>'
            f'<is><t xml:space="preserve">{txt}</t></is></c>')


# ---------------------------------------------------------
# mapeo hoja -> sheetN.xml (dinámico, por si cambia el orden)
# ---------------------------------------------------------

def _mapa_hojas(zf):
    wb = zf.read("xl/workbook.xml").decode("utf-8")
    rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    root = ET.fromstring(wb)
    sheets = []
    for s in root.find(f"{{{_NS_MAIN}}}sheets"):
        rid = s.get(f"{{{_NS_R}}}id")
        sheets.append((s.get("name"), rid))
    rroot = ET.fromstring(rels)
    rmap = {rel.get("Id"): rel.get("Target") for rel in rroot}
    out = {}
    for nombre, rid in sheets:
        target = rmap.get(rid, "")
        if not target.startswith("xl/"):
            target = "xl/" + target
        out[nombre] = target
    return out


# ---------------------------------------------------------
# estilos por columna (leídos de la fila 2 del sheet original)
# ---------------------------------------------------------

def _estilos_por_columna(sheet_xml):
    fila2 = re.search(r'<row r="2"[^>]*>(.*?)</row>', sheet_xml, re.DOTALL)
    estilos = {}
    if fila2:
        for col, s in re.findall(r'<c r="([A-Z]+)2"(?:\s+s="(\d+)")?', fila2.group(1)):
            estilos[col] = s or ""
    return estilos


def _cargar_shared_strings(zf):
    """Lista de textos de sharedStrings.xml (para resolver celdas
    tipo t='s' que referencian por índice)."""
    try:
        x = zf.read("xl/sharedStrings.xml").decode("utf-8")
    except KeyError:
        return []
    textos = []
    for si in re.findall(r"<si>(.*?)</si>", x, re.DOTALL):
        # un <si> puede tener varios <t> (texto con formato); se concatenan
        partes = re.findall(r"<t[^>]*>(.*?)</t>", si, re.DOTALL)
        txt = "".join(partes)
        txt = (txt.replace("&amp;", "&").replace("&lt;", "<")
               .replace("&gt;", ">"))
        textos.append(txt)
    return textos


def _headers_originales(sheet_xml, shared):
    """Encabezados (fila 1) en orden de columna, resolviendo tanto
    inlineStr como sharedString (t='s')."""
    fila1 = re.search(r'<row r="1"[^>]*>(.*?)</row>', sheet_xml, re.DOTALL)
    if not fila1:
        return []
    headers = []
    for m in re.finditer(r'<c r="([A-Z]+)1"([^>]*)>(.*?)</c>', fila1.group(1), re.DOTALL):
        attrs, contenido = m.group(2), m.group(3)
        if 't="s"' in attrs:
            vi = re.search(r"<v>(\d+)</v>", contenido)
            headers.append(shared[int(vi.group(1))] if vi else "")
        else:
            t = re.search(r"<t[^>]*>(.*?)</t>", contenido, re.DOTALL)
            if t:
                headers.append(t.group(1).replace("&amp;", "&")
                               .replace("&lt;", "<").replace("&gt;", ">"))
            else:
                headers.append("")
    return headers


# ---------------------------------------------------------
# construir el sheetData nuevo
# ---------------------------------------------------------

def _sheet_data_xml(headers, df, estilos):
    filas = []
    # fila 1: encabezados
    celdas = []
    for j, h in enumerate(headers):
        col = get_column_letter(j + 1)
        celdas.append(_celda_xml(f"{col}1", h, estilos.get(col, "")))
    filas.append(f'<row r="1">{"".join(celdas)}</row>')
    # datos
    for i in range(len(df)):
        r = df.iloc[i]
        celdas = []
        for j, h in enumerate(headers):
            col = get_column_letter(j + 1)
            celdas.append(_celda_xml(f"{col}{i + 2}", r[h], estilos.get(col, "")))
        filas.append(f'<row r="{i + 2}">{"".join(celdas)}</row>')
    return "<sheetData>" + "".join(filas) + "</sheetData>"


def _alinear(df_crudo, headers):
    """Devuelve un df con EXACTAMENTE las columnas 'headers' (orden
    de la plantilla). Empareja por nombre ignorando espacios/mayús;
    columnas faltantes quedan vacías."""
    norm = {str(c).strip().lower(): c for c in df_crudo.columns}
    salida = pd.DataFrame()
    for h in headers:
        real = norm.get(str(h).strip().lower())
        salida[h] = df_crudo[real] if real is not None else None
    return salida


def _actualizar_sheet(sheet_xml, headers, df, estilos):
    sd = _sheet_data_xml(headers, df, estilos)
    n_filas = len(df) + 1
    ultima = f"{get_column_letter(len(headers))}{n_filas}"
    sheet_xml = re.sub(r'<dimension ref="[^"]*"/>',
                       f'<dimension ref="A1:{ultima}"/>', sheet_xml, count=1)
    # reemplazo LITERAL (evita que \g, \1, etc. del contenido se interpreten)
    if "<sheetData>" in sheet_xml:
        ini = sheet_xml.index("<sheetData>")
        fin = sheet_xml.index("</sheetData>") + len("</sheetData>")
        sheet_xml = sheet_xml[:ini] + sd + sheet_xml[fin:]
    elif "<sheetData/>" in sheet_xml:
        sheet_xml = sheet_xml.replace("<sheetData/>", sd, 1)
    return sheet_xml, n_filas


def _solo_visible(workbook_xml, hoja_visible):
    """Deja visible SOLO la hoja indicada; oculta todas las demás.
    Ajusta activeTab a la hoja visible. No borra nada (seguro)."""
    sheets = list(re.finditer(r'<sheet\b[^>]*/>', workbook_xml))
    viejos = [m.group(0) for m in sheets]
    nuevas = []
    idx_visible = 0
    for i, tag in enumerate(viejos):
        nm = re.search(r'name="([^"]+)"', tag)
        nombre = nm.group(1) if nm else ""
        tag_sin = re.sub(r'\s+state="[^"]*"', "", tag)
        if nombre == hoja_visible:
            idx_visible = i
            nuevas.append(tag_sin)  # visible
        else:
            nuevas.append(tag_sin[:-2] + ' state="hidden"/>')
    for viejo, nuevo in zip(viejos, nuevas):
        workbook_xml = workbook_xml.replace(viejo, nuevo, 1)

    # activeTab -> índice de la hoja visible
    def _fix(m):
        wv = m.group(0)
        wv = re.sub(r'\s+activeTab="[^"]*"', "", wv)
        if wv.endswith("/>"):
            return wv[:-2] + f' activeTab="{idx_visible}"/>'
        return wv[:-1] + f' activeTab="{idx_visible}">'
    workbook_xml = re.sub(r'<workbookView\b[^>]*/?>', _fix,
                          workbook_xml, count=1)
    return workbook_xml


def _actualizar_pivotcache(cache_xml, hoja, n_filas):
    """Ajusta el ref del worksheetSource de esa hoja al nuevo número
    de filas y activa refreshOnLoad."""
    def _rep(m):
        attrs = m.group(0)
        if f'sheet="{hoja}"' not in attrs:
            return attrs
        # ref="A1:N684"  ->  ref="A1:N{n_filas}"
        return re.sub(r'ref="([A-Z]+1:[A-Z]+)\d+"',
                      lambda mm: f'ref="{mm.group(1)}{n_filas}"', attrs)
    cache_xml = re.sub(r'<worksheetSource[^>]*/>', _rep, cache_xml)
    if "refreshOnLoad" in cache_xml:
        cache_xml = re.sub(r'refreshOnLoad="[01]"', 'refreshOnLoad="1"', cache_xml)
    else:
        cache_xml = cache_xml.replace("<pivotCacheDefinition ",
                                      '<pivotCacheDefinition refreshOnLoad="1" ', 1)
    return cache_xml


# ---------------------------------------------------------
# función principal
# ---------------------------------------------------------

def generar_reporte(plantilla_path, solo_modulo=None):
    """Devuelve los BYTES del xlsx con las BD reemplazadas por lo
    guardado en Supabase (BD cruda). Las dinámicas quedan vivas.

    Si solo_modulo se indica (ventas/ingresos/cartera/saldo_proveedor),
    el archivo se descarga con SOLO la pestaña de ese módulo visible
    (las demás hojas se ocultan, no se borran, para no romper nada)."""
    with zipfile.ZipFile(plantilla_path) as z:
        nombres = z.namelist()
        contenido = {n: z.read(n) for n in nombres}
        shared = _cargar_shared_strings(z)

    mapa = _mapa_hojas(zipfile.ZipFile(plantilla_path))

    # filas nuevas por hoja (para ajustar los pivotCache)
    filas_por_hoja = {}

    for modulo, hoja in MODULO_HOJA.items():
        parte = mapa.get(hoja)
        if not parte or parte not in contenido:
            continue
        try:
            df_crudo = db.leer_crudo(modulo)
        except Exception:
            df_crudo = None
        if df_crudo is None or len(df_crudo) == 0:
            continue  # sin BD cruda -> se deja la hoja de la plantilla

        sheet_xml = contenido[parte].decode("utf-8")
        headers = _headers_originales(sheet_xml, shared)
        if not headers:
            continue
        estilos = _estilos_por_columna(sheet_xml)
        df_align = _alinear(df_crudo, headers)
        nuevo_xml, n_filas = _actualizar_sheet(sheet_xml, headers, df_align, estilos)
        contenido[parte] = nuevo_xml.encode("utf-8")
        filas_por_hoja[hoja] = n_filas

    # ajustar rangos de los pivotCache + refreshOnLoad
    for n in list(contenido.keys()):
        if re.search(r"xl/pivotCache/pivotCacheDefinition\d+\.xml$", n):
            x = contenido[n].decode("utf-8")
            for hoja, nf in filas_por_hoja.items():
                x = _actualizar_pivotcache(x, hoja, nf)
            # si la hoja de este cache no cambió, igual activamos refresh
            if "refreshOnLoad" in x:
                x = re.sub(r'refreshOnLoad="[01]"', 'refreshOnLoad="1"', x)
            else:
                x = x.replace("<pivotCacheDefinition ",
                              '<pivotCacheDefinition refreshOnLoad="1" ', 1)
            contenido[n] = x.encode("utf-8")

    # si se pidió solo un módulo, ocultar las demás hojas
    if solo_modulo and solo_modulo in MODULO_PIVOTE:
        hoja_vis = MODULO_PIVOTE[solo_modulo]
        wbxml = contenido["xl/workbook.xml"].decode("utf-8")
        wbxml = _solo_visible(wbxml, hoja_vis)
        contenido["xl/workbook.xml"] = wbxml.encode("utf-8")

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as z:
        for n in nombres:
            z.writestr(n, contenido[n])
    buffer.seek(0)
    return buffer.getvalue()